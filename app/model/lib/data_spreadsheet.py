from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

from app.model.lib.excel import export_to_xlsx

TIME_UNITS = {
    'h': 'hours',
    'm': 'minutes',
    's': 'seconds',
}

RED   = 'D02631'
WHITE = 'FFFFFF'

TECHNIQUE_DESCRIPTIONS = {
    'fc':          "Flow-cytometry values per time-point in {units}, use a period (.) as the decimal separator",
    'fc_ps':       "FC counts values per time-point for strain {strain} in {units}, use a period (.) as the decimal separator",
    'od':          "Optical density values per time-point, use a period (.) as the decimal separator",
    'plates':      "Plate count values per time-point, use a period (.) as the decimal separator",
    'plates_ps':   "Plate count values per time-point for strain {strain}, use a period (.) as the decimal separator.",
    '16s_ps':      "Abundances values per time-point for strain {strain}, use a period (.) as the decimal separator.",
    'ph':          "pH values per time-point, use a period (.) as the decimal separator",
    'metabolites': "Concentration values per time-point for metabolite {metabolite} in {units}, use a period (.) as the decimal separator",
    'STD':         "Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator",
}


def create_excel(submission, metabolite_names, strain_names):
    workbook = Workbook()

    short_time_units = submission.studyDesign['timeUnits']
    long_time_units = TIME_UNITS[short_time_units]

    headers_common = {
        # 'Position':           'Predetermine position based on the type of vessel specified before.',
        'Biological Replicate': 'Unique names of individual samples: a bottle, a well in a well-plate, or a mini-bioreactor.',
        'Compartment':          'A compartment within the vessel where growth is measured.',
        'Time':                 f"Measurement time-points in {long_time_units} ({short_time_units}).",
    }

    headers_bioreplicates = {**headers_common}
    headers_strains       = {**headers_common}
    headers_metabolites   = {**headers_common}

    for technique in submission.build_techniques():
        subject_type   = technique.subjectType
        technique_type = technique.type
        units          = technique.units

        if subject_type == 'bioreplicate':
            technique_name = technique.csv_column_name()
            description = TECHNIQUE_DESCRIPTIONS[technique_type].format(units=technique.units)

            headers_bioreplicates[technique_name] = description

            if technique.includeStd:
                title = ' '.join([technique_name, 'STD'])
                headers_bioreplicates[title] = TECHNIQUE_DESCRIPTIONS['STD']

        elif subject_type == 'strain':
            for strain_name in strain_names:
                title = technique.csv_column_name(strain_name)

                description = TECHNIQUE_DESCRIPTIONS[f"{technique_type}_ps"].format(
                    strain=strain_name,
                    units=units,
                )
                headers_strains[title] = description

                if technique.includeStd:
                    title = ' '.join([title, 'STD'])
                    headers_strains[title] = TECHNIQUE_DESCRIPTIONS['STD']

        elif subject_type == 'metabolite':
            for metabolite in metabolite_names:
                title = technique.csv_column_name(metabolite)
                description = TECHNIQUE_DESCRIPTIONS['metabolites'].format(
                    metabolite=metabolite,
                    units=units,
                )
                headers_metabolites[title] = description

                if technique.includeStd:
                    title = ' '.join([metabolite, 'STD'])
                    headers_metabolites[title] = TECHNIQUE_DESCRIPTIONS['STD']

        else:
            raise ValueError(f"Invalid technique subject_type: {subject_type}")

    # Create sheets for each category of measurement:
    if len(headers_bioreplicates) > 3:
        _fill_sheet(workbook, "Growth data per community", headers_bioreplicates, submission)

    if len(headers_strains) > 3:
        _fill_sheet(workbook, "Growth data per strain", headers_strains, submission)

    if len(headers_metabolites) > 3:
        _fill_sheet(workbook, "Growth data per metabolite", headers_metabolites, submission)

    return export_to_xlsx(workbook)


def _add_header(sheet, index, title, description, fill_color):
    cell         = sheet.cell(row=1, column=index, value=title)
    cell.comment = Comment(description, author="μGrowthDB")

    # Built-in styles:
    # <https://openpyxl.readthedocs.io/en/stable/styles.html#using-builtin-styles>
    #
    cell.style = 'Headline 3'

    # Calculate column width based on the length of the title text, plus some extra space
    column_width = len(title)
    column_letter = get_column_letter(index)
    sheet.column_dimensions[column_letter].width = column_width + 1


def _fill_sheet(workbook, sheet_title, headers, submission):
    if workbook.sheetnames == ['Sheet']:
        # Then we have a brand new workbook, use its first sheet
        sheet = workbook.active
        workbook.active.title = sheet_title
    else:
        sheet = workbook.create_sheet(title=sheet_title)

    # Add headers and descriptions to the first row and modify the width of each columns
    for index, (title, description) in enumerate(headers.items(), start=1):
        if title.endswith(' STD'):
            fill_color = WHITE
        else:
            fill_color = RED

        _add_header(sheet, index, title, description, fill_color=fill_color)

    bottom_border = Border(bottom=Side(style="thin", color="000000"))

    row_index = 2
    for experiment in submission.studyDesign['experiments']:
        for bioreplicate in experiment['bioreplicates']:
            for compartment_name in experiment['compartmentNames']:
                for _ in range(experiment['timepointCount']):
                    sheet.cell(row=row_index, column=1, value=bioreplicate['name'])
                    sheet.cell(row=row_index, column=2, value=compartment_name)

                    row_index += 1

                # Apply border after every group of time points
                for column in range(1, len(headers.keys()) + 1):
                    cell = sheet.cell(row=(row_index - 1), column=column)
                    cell.border = bottom_border

    # Freeze header and label columns
    sheet.freeze_panes = "D2"
