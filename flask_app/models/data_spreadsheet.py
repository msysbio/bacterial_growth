from tempfile import NamedTemporaryFile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

def create_excel(
        measure_options,
        meta_options,
        type_vessel,
        number_vessels,
        number_columns,
        number_rows,
        number_timepoints,
        taxa
    ):
    # Create a new workbook
    wb = Workbook()

    ws3 = wb.active
    wb.active.title = "Replicate_metadata"


    headers3_constant = {
        'Biological_Replicate_id': 'Unique IDs of individual samples: a bottle, a well in a well-plate or mini-bioreactor',
        'Biosample_link': 'Provide the Biosample link if available.',
        'Description' : 'Provide an informative description for the Biological_Replicate_ids.',
    }

    for col_idx, (header, description) in enumerate(headers3_constant.items(), start=1):
            cell = ws3.cell(row=1, column=col_idx, value=header)
            cell.comment = Comment(description, author="ExcelBot")
            cell.font = Font(bold=True)  # Make the header bold
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")

            # Calculate column width based on the length of the header text
            column_width = len(header)
            column_letter = get_column_letter(col_idx)
            ws3.column_dimensions[column_letter].width = column_width +  3 #Gives extra space

    # Select the active worksheet
    wb.create_sheet(title="Growth_Data_and_Metabolites")
    ws = wb["Growth_Data_and_Metabolites"]

    # Populate raw data template based on user's input
    positions = generate_positions(type_vessel,number_vessels,number_columns,number_rows,number_timepoints)

    headers1_constant = {
        'Position' : 'Predetermine position based on the type of vessel specified before.',
        'Biological_Replicate_id': 'Unique IDs of individual samples: a bottle, a well in a well-plate or mini-bioreactor',
        'Time' : 'Measurement time-points in the format: hh:mm:ss (hour, minutes and seconds)',
        'pH': 'pH values per time-point, use a period (.) as the decimal separator'
    }

    header1_OD = {
        'OD' : 'Optical density values per time-point, use a period (.) as the decimal separator',
        'OD_std' : 'Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator'

    }

    header1_plate_counts = {
        'Plate_counts' : 'Plate counts values per time-point, use a period (.) as the decimal separator',
        'Plate_counts_std' : 'Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator'

    }

    header1_FC = {
        'FC' : 'Flow-cytometry values per time-point, use a period (.) as the decimal separator',
        'FC_std' : 'Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator'

    }

    header1_color_white = ['pH','OD_std', 'Plate_counts_std','Biosample_link']

    if 'od' in measure_options:
        headers1_constant.update(header1_OD)

    if 'plates' in measure_options:
        headers1_constant.update(header1_plate_counts)

    if 'fc' in measure_options:
        header1_FC.update(header1_FC)


    # Add metabolites measured as column names
    for i in meta_options:
        header1_metabolites = {
            f'{i}' : 'Concentration values per time-point, use a period (.) as the decimal separator',
            f'{i}_std' : 'Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator'
        }
        headers1_constant.update(header1_metabolites)


    # Add headers and descriptions to the first row and modify the width of each columns
    for col_idx, (header, description) in enumerate(headers1_constant.items(), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.comment = Comment(description, author="ExcelBot")
        cell.font = Font(bold=True)  # Make the header bold

        if header in header1_color_white or header.endswith('std'):
            cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
        else:
            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")

        # Calculate column width based on the length of the header text
        column_width = len(header)
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = column_width +  3 #Gives extra space

    # Fill the position column with generated positions
    for idx, position in enumerate(positions, start=2):
        ws.cell(row=idx, column=1, value=position)


    headers2_constant = {
            'Position' : 'Predetermine position based on the type of vessel specified before.',
            'Biological_Replicate_id': 'Unique IDs of individual samples: a bottle, a well in a well-plate or mini-bioreactor',
            'Time' : 'Measurement time-points in the format: hh:mm:ss (hour, minutes and seconds)',
        }

    if 'rna' in measure_options:
        for i in taxa:
            header2_species = {
                f'{i}_reads' : f'Abundances values per time-point for species {i}, use a period (.) as the decimal separator',
                f'{i}_reads_std' : 'Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator'
            }
            headers2_constant.update(header2_species)

    if 'fc_ps' in measure_options:
        for i in taxa:
            header2_species_fc = {
                f'{i}_counts' : f'FC counts values per time-point for species {i}, use a period (.) as the decimal separator',
                f'{i}_counts_std' : 'Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator'

            }
            headers2_constant.update(header2_species_fc)

    if 'plates_ps' in measure_options:
        for i in taxa:
            header2_species_pc = {
                f'{i}_Plate_counts' : f'FC counts values per time-point for species {i}, use a period (.) as the decimal separator',
                f'{i}_Plate_counts_std' : 'Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator'

            }
            headers2_constant.update(header2_species_pc)

    if 'rna' or 'fc_ps' or 'plates_ps' in measure_options:
        wb.create_sheet(title="growth_per_species")
        ws2 = wb["growth_per_species"]

        for col_idx, (header, description) in enumerate(headers2_constant.items(), start=1):
            cell = ws2.cell(row=1, column=col_idx, value=header)
            cell.comment = Comment(description, author="ExcelBot")
            cell.font = Font(bold=True)  # Make the header bold

            if header.endswith('std'):
                cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
            else:
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type="solid")

            # Calculate column width based on the length of the header text
            column_width = len(header)
            column_letter = get_column_letter(col_idx)
            ws2.column_dimensions[column_letter].width = column_width +  3 #Gives extra space

        for idx, position in enumerate(positions, start=2):
            ws2.cell(row=idx, column=1, value=position)

    temp_file = NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()
    with open(temp_file.name, 'rb') as f:
        data_excel = f.read()
    return data_excel


def generate_positions(type_vessel, number_vessels, number_columns, number_rows, number_timepoints):
    positions = []
    if type_vessel == 'bottles' or type_vessel == 'agar_plates':
        if number_vessels and number_timepoints:
            for vessel_num in range(1, int(number_vessels)+ 1):
                for _ in range(int(number_timepoints)):
                    positions.append(f"{type_vessel}{vessel_num}")

    if type_vessel == 'well_plates' or type_vessel == 'mini_react':
        if number_columns and number_rows and number_timepoints:
            for row in range(1, int(number_rows) + 1):
                for col in range(1, int(number_columns) + 1):
                    for _ in range(int(number_timepoints)):
                            positions.append(f"{get_column_letter(col)}{row}")

    return positions
