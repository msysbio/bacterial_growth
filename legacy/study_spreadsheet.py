from tempfile import NamedTemporaryFile
import openpyxl

from lib.excel import export_to_xlsx


def create_excel(
    keywords,
    list_taxa_id,
    all_strain_data,
    create_private_project_id,
    unique_study_id_val,
):
    """
    keywords:        NCBI Taxonomy Ids of taxa with exact match to a NCBI Taxonomy Id
    list_taxa_id:    names of the taxa with exact NCBI Taxonomy ids
    all_strain_data: list of dictionaries with taxa without an exact NCBI Taxonomy Id
    """

    workbook = openpyxl.load_workbook('templates/excel/metadata_template.xlsx')
    sheet = workbook['COMMUNITY_MEMBERS']

    k = 0  # Initialize k here

    if keywords:
        for k in range(1, len(keywords) + 1):
            sheet.cell(row=k+1, column=1, value=f'member_{k}')
            sheet.cell(row=k+1, column=2, value=1)
            sheet.cell(row=k+1, column=3, value=keywords[k-1])
            sheet.cell(row=k+1, column=4, value=list_taxa_id[k-1])

    if all_strain_data:
        for i, dic in enumerate(all_strain_data):
            name = dic.get('name', '')

            if name is None or name == "":
                continue

            description = dic.get('description', '')
            parent_ncbi_tax_id = dic.get('parent_taxon_id')

            sheet.cell(row=k+i+2, column=1, value=f'member_{k+i+1}')
            sheet.cell(row=k+i+2, column=2, value=0)
            sheet.cell(row=k+i+2, column=3, value=name)
            sheet.cell(row=k+i+2, column=4, value=parent_ncbi_tax_id)
            # TODO (2025-03-12) Assembly ID in "new strain" form:
            sheet.cell(row=k+i+2, column=5, value='')
            sheet.cell(row=k+i+2, column=6, value=description)

    sheet_study = workbook['STUDY']
    if create_private_project_id is not None:
        sheet_study.cell(row=2, column=1, value=create_private_project_id)
    if unique_study_id_val is not None:
        sheet_study.cell(row=2, column=2, value=unique_study_id_val)

    return export_to_xlsx(workbook)
