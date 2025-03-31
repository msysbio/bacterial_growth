from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill

from lib.excel import export_to_xlsx

TIME_UNITS = {
    'd': 'days',
    'h': 'hours',
    'm': 'minutes',
    's': 'seconds',
}

TECHNIQUE_NAMES = {
    'ph':     'pH',
    'fc':     'FC',
    'od':     'OD',
    'plates': 'PC',
    '16s':    '16S-rRNA reads',
}

RED   = 'D02631'
WHITE = 'FFFFFF'

TECHNIQUE_DESCRIPTIONS = {
    'fc':          "Flow-cytometry values per time-point in {units}, use a period (.) as the decimal separator",
    'fc_ps':       "FC counts values per time-point for strain {strain} in {units}, use a period (.) as the decimal separator",
    'od':          "Optical density values per time-point, use a period (.) as the decimal separator",
    'plates':      "Plate count values per time-point, use a period (.) as the decimal separator",
    'plates_ps':   "Plate count values per time-point for strain {strain}, use a period (.) as the decimal separator.",
    '16s_ps':      "Abundances values per time-point for strain {strain}, use a period (.) as the decimal separator.",
    'ph':          "pH values per time-point, use a period (.) as the decimal separator",
    'metabolites': "Concentration values per time-point for metabolite {metabolite} in {units}, use a period (.) as the decimal separator",
    'STD':         "Standard deviation if more than 1 technical replicate, use a period (.) as the decimal separator",
}


def create_excel(submission, metabolite_names, strain_names):
    workbook = Workbook()

    metadata_sheet = workbook.active
    workbook.active.title = "Replicate_metadata"

    headers_info = {
        'Biological_Replicate_id': 'Pick unique identifiers of individual samples: a bottle, a well in a well-plate or mini-bioreactor',
        'Biosample_link': 'Provide the biosample link if available.',
        'Description': 'Provide an informative description for the biological replicate',
    }

    for index, (title, description) in enumerate(headers_info.items(), start=1):
        _add_header(metadata_sheet, index, title, description, fill_color=RED)

    # Populate raw data template based on user's input
    positions = _generate_positions(
        submission.studyDesign['vessel_type'],
        submission.studyDesign['vessel_count'],
        submission.studyDesign['column_count'],
        submission.studyDesign['row_count'],
        submission.studyDesign['timepoint_count'],
    )

    short_time_units = submission.studyDesign['time_units']
    long_time_units = TIME_UNITS[short_time_units]

    headers_common = {
        'Position':                'Predetermine position based on the type of vessel specified before.',
        'Biological_Replicate_id': 'Unique IDs of individual samples: a bottle, a well in a well-plate or mini-bioreactor.',
        'Time':                    f"Measurement time-points in {long_time_units} ({short_time_units}).",
    }

    headers_bioreplicates = {**headers_common}
    headers_strains       = {**headers_common}
    headers_metabolites   = {**headers_common}

    for technique in submission.studyDesign['techniques']:
        subject_type   = technique['subjectType']
        technique_type = technique['type']
        units          = technique['units']

        if subject_type == 'bioreplicate':
            technique_name = TECHNIQUE_NAMES[technique_type]
            description = TECHNIQUE_DESCRIPTIONS[technique_type].format(units=technique['units'])

            headers_bioreplicates[technique_name] = description

        elif subject_type == 'strain':
            technique_name = TECHNIQUE_NAMES[technique_type]

            for strain_name in strain_names:
                if technique_type in ('16s', 'plates'):
                    title = ' '.join([strain_name, units])
                else:
                    title = ' '.join([strain_name, technique_name, f"({units})"])

                description = TECHNIQUE_DESCRIPTIONS[f"{technique_type}_ps"].format(
                    strain=strain_name,
                    units=units,
                )
                headers_strains[title] = description

                if technique['includeStd']:
                    title = ' '.join([strain_name, technique_name, 'STD'])
                    headers_strains[title] = TECHNIQUE_DESCRIPTIONS['STD']

        elif subject_type == 'metabolite':
            for metabolite in metabolite_names:
                title = ' '.join([metabolite, f"({units})"])
                description = TECHNIQUE_DESCRIPTIONS['metabolites'].format(
                    metabolite=metabolite,
                    units=units,
                )
                headers_metabolites[title] = description

                if technique['includeStd']:
                    title = ' '.join([metabolite, 'STD'])
                    headers_metabolites[title] = TECHNIQUE_DESCRIPTIONS['STD']

        else:
            raise ValueError(f"Invalid technique subject_type: {subject_type}")

    # Create sheets for each category of measurement:
    if len(headers_bioreplicates) > 3:
        _fill_sheet(workbook, "Growth data per bioreplicate", headers_bioreplicates, positions)

    if len(headers_strains) > 3:
        _fill_sheet(workbook, "Growth data per strain", headers_strains, positions)

    if len(headers_metabolites) > 3:
        _fill_sheet(workbook, "Growth data per metabolite", headers_metabolites, positions)

    return export_to_xlsx(workbook)


def _add_header(sheet, index, title, description, fill_color):
    cell         = sheet.cell(row=1, column=index, value=title)
    cell.comment = Comment(description, author="μGrowthDB")
    cell.font    = Font(bold=True)
    cell.fill    = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    # Calculate column width based on the length of the title text, plus some extra space
    column_width = len(title)
    column_letter = get_column_letter(index)
    sheet.column_dimensions[column_letter].width = column_width + 1


def _fill_sheet(workbook, sheet_title, headers, positions):
    sheet = workbook.create_sheet(title=sheet_title)

    # Add headers and descriptions to the first row and modify the width of each columns
    for index, (title, description) in enumerate(headers.items(), start=1):
        if title.endswith(' STD'):
            fill_color = WHITE
        else:
            fill_color = RED

        _add_header(sheet, index, title, description, fill_color=fill_color)

    # Fill the position column with generated positions
    for idx, position in enumerate(positions, start=2):
        sheet.cell(row=idx, column=1, value=position)


def _generate_positions(type_vessel, number_vessels, number_columns, number_rows, number_timepoints):
    positions = []
    if type_vessel == 'bottles' or type_vessel == 'agar_plates':
        if number_vessels and number_timepoints:
            for vessel_num in range(1, int(number_vessels) + 1):
                for _ in range(int(number_timepoints)):
                    positions.append(f"{type_vessel}{vessel_num}")

    if type_vessel == 'well_plates' or type_vessel == 'mini_react':
        if number_columns and number_rows and number_timepoints:
            for row in range(1, int(number_rows) + 1):
                for col in range(1, int(number_columns) + 1):
                    for _ in range(int(number_timepoints)):
                        positions.append(f"{get_column_letter(col)}{row}")

    return positions
