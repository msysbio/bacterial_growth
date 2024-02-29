import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from constants import LOCAL_DIRECTORY


# Create a new workbook
wb = Workbook()

# Select the active worksheet
ws = wb.active

wb.active.title = "README"
wb.create_sheet(title="STUDY_DATA")
wb.create_sheet(title="COMUNITY_MEMBERS")

data_readme = {
    'HOW TO COMPLETE': "In this excel you will find 3 sheets: README, STUDY_DATA and COMUNITY_MEMBERS."
    " First, read CAREFULLY this README. Complete each field with the correct format, numerical or textual"
    " depending on the column. ONLY add spaces in the fields where this is explicitly specified.",
    'STUDY_DATA':"In this sheet you will complete all the details of the study. Each column represent"
    " a specific information and each row represents a BIOLOGIAL REPLICATE. Fill as many rows as many"
    " Biological replicates you have in your study. Bellow each column is explained.",
    'Study_ID': "This corresponds to the unique ID given before downloading this file."
    " Do NOT add spaces or any type of special characters.",
    'Study_Name' : "This corresponds to the name of your study. Give an informative and name. Adding spaces is allowed.",
    'Study_Description': "Complete with a clear an concise description. Adding spaces is allowed.",
    'Study_PublicationURL': "If this study has been published fill in with the publication URL. If NOT just leave it blank.",
    'BiologicalReplicate_Name': "Name of the biological replicate. Adding spaces is allowed",
    'BiologicalReplicate_Description': "Description of the biological replicate. Adding spaces is allowed.",
    'Reactor_Name': "Name of the reactor used for the biological replicate. Adding spaces is allowed.",
    'Reactor_Description': "Description of the reactor used for the biological replicate. Adding spaces is allowed.",
    'Reactor_Volume':"Fill in with the Volume used for the reactor. Volume in mL. Numerical values ONLY.",
    'Reactor_Atmosphere': "Atmosphere in XX. Numerical values ONLY.",
    'Reactor_StirringSpeed': "Stirring speed used in rpm. Numerical values ONLY.",
    'Reactor_Mode': 'Complete with the following options: chemostat/batch/fed-batch. Do NOT use any other options. Do NOT add spaces.',
    'Reactor_ContainersNumber' : 'Complete with the number of containers. Numerical values ONLY.',
    'Media_Name': 'Name of the media used in the biological replicate. Adding spaces is allowed.',
    'Media_Path': 'Give the complete path to the file that describe your media.',
    'Media_Recipe': 'Provide the link of the media used if available.',
    'Comunity_Groups':'For each row specify with a letter the distinct bacterial strains used in each bioloical replicate.'
    ' For example, if you used two for a given biological replicate, Strain 1 will be A and Strain 2 will be B. Fill in as: A,B'
    ' IF in another biological replicate you use only Strain 1, fill in the second row as: A. Do NOT add spaces.',
    'Blank':'Fill in with 1 if the biological replicate is blank. 0 otherwise. DO NOT leave in blank.',
    'Initial_pH' : 'Complete with the initial pH used. Numerical values ONLY.',
    'Initial_Temperature': 'Initial temperature in Celsius. Numerical values ONLY.',
    'Carbon_Sourced': 'Fill in with 1 if carbon sourced. 0 otherwise. DO NOT leave in blank.',
    'Antibiotic': 'Fill in with 1 if antibiotic present. 0 otherwise. DO NOT leave in blank.',
    'Plate_Number': 'Number of plate. Numerical values ONLY.',
    'Plate_Position': 'List of locations in the plate, i.e: 1A, 1B, 2C.',
    'Files_Directory' : 'Comma separated list of directories. CAREFULL! Keep the order as the plate positions.',
    'Perturbation_Property' : 'If your biological replicate had perturbations, complete each column starting with Perturbation.'
    ' Leave all columns in BLANK if no perturbations present. Indicate which property has been perturbed.',
    'Perturbation_Description' : 'Description of the perturbation.Adding spaces allowed.',
    'Perturbation_NewValue' : 'New value of the perturbed property. Numerical values ONLY.',
    'Perturbation_StartTime' : 'Time in minutes. Numerical values ONLY.',
    'Perturbation_EndTime':'Time in minutes.Numerical values ONLY.',
    'Perturbation_Plate_Number' : 'Number of plate perturbed. Numerical values ONLY.',
    'Perturbation_Plate_Position' :'List of locations in the plate, i.e: 1A, 1B, 2C',
    'Perturbation_FilesDirectory': 'Comma separated list of directories. CAREFULL! Keep the order as the plate positions.',
    'COMUNITY_MEMBERS':'In this sheet you will complete all information related with the comunities or bacteria strains used in the study.'
    ' Every row corresponds to a DIFFERENT bacterial strain NOT to a biological replicate.',
    'Comunity_ID': 'Complete ONLY with the same letters used in Comunity_Groups. One letter per row. Following the example provided in Comunity_Groups'
    ' There will be two rows: A and B',
    'Bacteria_Genus':'Give the exact Genus of your bacteria strain.',
    'Bacteria_Species': 'Give the exact Species of your bacterial strain.',
    'Bacteria_NCBISpeciesID': 'Give the species NCBI ID of your bacterial strain.',
    'Bacteria_Strain':'Give the exact strain name of your bacteria.',
    'Bacteria_NCBIStrainID': 'Give the strai NCBI ID of your bacteria.',
    'Bacteria_GeneticallyModified': 'Fill in with 1 if mutant, 0 Otherwise.',
    }

df_data_readme = pd.DataFrame([data_readme],columns=data_readme.keys())

for row_idx, (title, values) in enumerate(data_readme.items(), start=1):
    cell1 = ws.cell(row=row_idx, column=1, value=title)
    cell2 = ws.cell(row=row_idx, column=2, value=values)
    cell1.font = Font(bold=True)
    cell1.alignment = Alignment(wrapText=False,vertical='center')
    cell2.alignment = Alignment(wrapText=True,vertical='top')

ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 80


ws2 = wb["STUDY_DATA"]
ws3 = wb["COMUNITY_MEMBERS"]


# Add headers to the columns
headers_1 = {
    'Study_ID': 'Complete with the ID provided before for your study. Do NOT add spaces.',
    'Study_Name' : 'Give an informative name to your study.',
    'Study_Description': 'Complete with a clear an concise description.',
    'Study_PublicationURL': 'Fill in with the publication URL if published.',
    'BiologicalReplicate_Name': 'Name of the biological replicate. Each row is a DIFFERENT biological replicate.',
    'BiologicalReplicate_Description': 'Description of the biological replicate.',
    'Reactor_Name': 'Name of the reactor.',
    'Reactor_Description': 'Description of the reactor.',
    'Reactor_Volume':'Volume in mL.',
    'Reactor_Atmosphere': 'Atmosphere in XX.',
    'Reactor_StirringSpeed': 'Stirring speed used in rpm.',
    'Reactor_Mode': 'Complete with the following options: chemostat/batch/fed-batch.',
    'Reactor_ContainersNumber' : 'Complete with the number of containers.',
    'Media_Name': 'Name of the media.',
    'Media_Path': 'Give the complete path to the file that describe your media.',
    'Media_Recipe': 'Provide the link if available.',
    'Comunity_Groups':'For each row specify with a letter the bacterial strains.',
    'Blank':'Fill in with 1 if the biological replicate is blank. 0 otherwise. DO NOT leave in blank.',
    'Inoculum_Concentration': 'Indicated in cells per mL.',
    'Inoculum_Volume': 'Indicated in mL.',
    'Initial_pH' : 'Complete with the initial pH.',
    'Initial_Temperature': 'Initial temperature in Celsius.',
    'Carbon_Sourced': 'Fill in with 1 if carbon sourced. 0 otherwise. DO NOT leave in blank.',
    'Antibiotic': 'Fill in with 1 if antibiotic present. 0 otherwise. DO NOT leave in blank.',
    'Plate_Number': 'Number of plate.',
    'Plate_Position': 'List of locations in the plate, i.e: 1A, 1B, 2C.',
    'Files_Directory' : 'Comma separated list of directories. Keep the order as the plate positions.',
    'Perturbation_Property' : 'Indicate which property has been perturbed.',
    'Perturbation_Description' : 'Description of the perturbation.',
    'Perturbation_NewValue' : 'New value of the perturbed property.',
    'Perturbation_StartTime' : 'Time in minutes.',
    'Perturbation_EndTime':'Time in minutes.',
    'Perturbation_Plate_Number' : 'Number of plate perturbed',
    'Perturbation_Plate_Position' :'List of locations in the plate, i.e: 1A, 1B, 2C',
    'Perturbation_FilesDirectory': 'Comma separated list of directories. Keep the order as the plate positions.'
}

headers_2 = {
    'Comunity_ID': 'Use the same letters used in Comunity_Groups',
    'Bacteria_Genus':'Give the exact Genus of your bacteria strain.',
    'Bacteria_Species': 'Give the exact Species of your bacterial strain.',
    'Bacteria_NCBISpeciesID': 'Give the species NCBI ID of your bacterial strain.',
    'Bacteria_Strain':'Give the exact strain name of your bacteria.',
    'Bacteria_NCBIStrainID': 'Give the strai NCBI ID of your bacteria.',
    'Bacteria_GeneticallyModified': 'Fill in with 1 if mutant, 0 Otherwise.',
}

# Add headers and descriptions to the first row and modify the width of each columns
for col_idx, (header, description) in enumerate(headers_1.items(), start=1):
    cell = ws2.cell(row=1, column=col_idx, value=header)
    cell.comment = Comment(description, author="ExcelBot")
    cell.font = Font(bold=True)  # Make the header bold

    # Calculate column width based on the length of the header text
    column_width = len(header)
    column_letter = get_column_letter(col_idx)
    ws2.column_dimensions[column_letter].width = column_width +  3 #Gives extra space

# Add headers and descriptions to the first row and modify the width of each columns
for col_idx, (header, description) in enumerate(headers_2.items(), start=1):
    cell = ws3.cell(row=1, column=col_idx, value=header)
    cell.comment = Comment(description, author="ExcelBot")
    cell.font = Font(bold=True)  # Make the header bold

    # Calculate column width based on the length of the header text
    column_width = len(header)
    column_letter = get_column_letter(col_idx)
    ws3.column_dimensions[column_letter].width = column_width +  3 #Gives extra space

# saves excel file in the locl directory
template_filename = LOCAL_DIRECTORY + 'submission_excel_template.xlsx'
wb.save(template_filename)
